import tkinter as tk
from tkinter import messagebox, filedialog
from tkcalendar import Calendar, DateEntry
from datetime import datetime, date
from openpyxl import load_workbook, Workbook
#COLORES
GRIS = "#d8cccc"
NEGRO = "#111111"

def Cargar_datos_Excel(archivo):# carga datos del excel
    wb = load_workbook(archivo)
    ws = wb.active
    registros = []
    for row in ws.iter_rows(min_row=2): #itera las filas del excel
        if len(row) == 5:
            distribuidora_Name , cargador, distco, volumen, fecha = row
            registros.append((distribuidora_Name, cargador, distco, volumen, fecha))
        else:
            messagebox.showerror("ERROR","Error en el registro de datos")
    
    return registros

def Guarda_Datos_En_Excel(records, cliente_promedio):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Promedios"
    ws["A1"] = "Distribuidora"
    ws["B1"] = "Fecha"
    ws["C1"] = "Distco"
    ws["D1"] = "Volumen Promedio"

    row = 2 #contador

    for record in records:
        ws[f"A{row}"] = record[0].value
        ws[f"B{row}"] = record[4].value
        ws[f"C{row}"] = record[2].value
        ws[f"D{row}"] = cliente_promedio[record[0].value]
        row += 1
    
    wb.save(f"Promedios.xlsx")
    messagebox.showinfo("EXITO","El archivo fue creado con exito.")


    
def Cargar_Datos():#seleccionar el archivo a cargar
    archivo = filedialog.askopenfilename(title="Seleccionar archivo de Excel", filetypes=(("Archivos de Excel", "*.xlsx"), ("Todos los archivos", "*.*")))
    if archivo:
        global registros
        registros = Cargar_datos_Excel(archivo)
        messagebox.showinfo("Éxito", "Los datos fueron cargados")

def Calcular_Promedio_Individual(records):
    client_volumen = {}
    for record in records:
        client = record[0].value
        if client not in client_volumen:
            client_volumen[client] = []
        client_volumen[client].append(record[3].value)
        
    cliente_promedio = {client: sum(volumen) / len(volumen) for client, volumen in client_volumen.items()}
    return cliente_promedio

def Calcular_Volumen_Promedio(records):#calcula el promedio total
    volumen_Total = sum(record[3].value for record in records)
    Volumen_Promedio = volumen_Total / len(records)
    return Volumen_Promedio

def Calcular_Promedios():
    for widget in root.winfo_children(): #limpiar la screen cada vez que llama la funcion
        widget.destroy()
        
    lbl_Fecha_Incio = tk.Label(root, text="Fecha de Inicio:", bg="gray",fg=NEGRO)
    lbl_Fecha_Incio.pack(pady=10)
    #entry de la fecha
    entry_Fecha_Inicio = DateEntry(root, date_pattern='yyyy-mm-dd')
    entry_Fecha_Inicio.pack(pady=10)

    def Enviar_Promedio():
        fecha_inicio = entry_Fecha_Inicio.get_date()
        print(f"{fecha_inicio}")# mausqueherramienta
        parametros = []
        for record in registros:
            if len(record) == 5:
                fecha_record = record[4].value
                if isinstance(fecha_record, str):
                    try:
                        fecha_record = datetime.strptime(fecha_record, "%Y/%m/%d").date()
                    except ValueError:
                        continue
                
                if isinstance(fecha_record, datetime):
                    fecha_record = fecha_record.date()
                    
                if fecha_record >= fecha_inicio:
                    parametros.append(record)
        
        if not parametros:
            messagebox.showerror("Sin datos", "No hay registros desde la fecha seleccionada.")
            return
        volumen_Promedio = Calcular_Volumen_Promedio(parametros)
        cliente_promedio = Calcular_Promedio_Individual(parametros)
        Guarda_Datos_En_Excel(parametros,cliente_promedio)
        messagebox.showinfo("Promedio de Volumen", f"El promedio de volumen desde {fecha_inicio} es {volumen_Promedio}")

    btn_Calcular_Promedio = tk.Button(root,text="Calcular", bg=GRIS,fg=NEGRO, command=Enviar_Promedio) #btn para calcular el promedio
    btn_Calcular_Promedio.pack(pady=10)

    btn_Back_Menu  = tk.Button(root,text="VOLVER",bg=GRIS,fg=NEGRO, command=Main_Menu)
    btn_Back_Menu.pack(pady=10)

def Main_Menu():#funciones del menu principal
    for widget in root.winfo_children(): #limpiar la screen cada vez que llama la funcion
        widget.destroy()
    #Botones del meno principal
    btn_Cargar = tk.Button(root, text="Cargar Datos", bg=GRIS,fg=NEGRO,width=20,height=3, command=Cargar_Datos)
    btn_Cargar.pack(side=tk.LEFT, padx=10,pady=10)

    btn__Ver_Registros = tk.Button(root,text="Ver Registros", bg=GRIS,fg=NEGRO,width=20,height=3)
    btn__Ver_Registros.pack(side=tk.LEFT, padx=10, pady=10)

    btn_Calcular_Promedio = tk.Button(root,text="Calcular Promedio", bg=GRIS,fg=NEGRO,width=20,height=3,command=Calcular_Promedios)
    btn_Calcular_Promedio.pack(side=tk.LEFT,padx=10,pady=10)

#ventana settings
root = tk.Tk()
root.title("Distribuidoras Mabel")
root.geometry("720x460")
root.config(bg="GRAY")

registros = []#para almacenar los registros

Main_Menu()

root.mainloop()